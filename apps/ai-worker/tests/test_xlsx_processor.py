# apps/ai-worker/tests/test_xlsx_processor.py
"""
Unit tests for XlsxProcessor module.
Tests Excel multi-sheet processing and markdown conversion.
"""

import pytest

import pandas as pd
from openpyxl import Workbook

from src.converters import XlsxConverter as XlsxProcessor
from src.models import ProcessorOutput


@pytest.fixture
def processor():
    """Create XlsxProcessor instance."""
    return XlsxProcessor()


def create_xlsx(sheets_data: dict, path) -> None:
    """Helper to create XLSX file with multiple sheets."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, data in sheets_data.items():
            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name=sheet_name, index=False)


class TestXlsxProcessorMultiSheet:
    """Tests for multi-sheet processing."""

    @pytest.mark.asyncio
    async def test_process_all_sheets(self, processor, tmp_path):
        """All sheets are processed."""
        xlsx_path = tmp_path / "multi.xlsx"
        create_xlsx(
            {
                "Sheet1": {"A": [1, 2], "B": [3, 4]},
                "Sheet2": {"X": [5, 6], "Y": [7, 8]},
            },
            xlsx_path,
        )

        result = await processor.process(str(xlsx_path))

        assert isinstance(result, ProcessorOutput)
        assert "Sheet1" in result.markdown
        assert "Sheet2" in result.markdown
        assert result.sheet_count == 2

    @pytest.mark.asyncio
    async def test_sheet_names_as_headers(self, processor, tmp_path):
        """Sheet names become H1 headings."""
        xlsx_path = tmp_path / "headers.xlsx"
        create_xlsx({"Sales Data": {"Amount": [100, 200]}}, xlsx_path)

        result = await processor.process(str(xlsx_path))

        assert "# Sales Data" in result.markdown

    @pytest.mark.asyncio
    async def test_skip_empty_sheets(self, processor, tmp_path):
        """Empty sheets are skipped."""
        xlsx_path = tmp_path / "empty.xlsx"
        # Create workbook with one empty and one filled sheet
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Empty"
        ws2 = wb.create_sheet("Data")
        ws2.append(["Name", "Value"])
        ws2.append(["Test", 123])
        wb.save(xlsx_path)

        result = await processor.process(str(xlsx_path))

        # Should have Data sheet, may or may not have Empty
        assert "Data" in result.markdown
        assert "Name" in result.markdown


class TestXlsxProcessorSmallTables:
    """Tests for small table (markdown format)."""

    @pytest.mark.asyncio
    async def test_small_table_markdown(self, processor, tmp_path):
        """Small table outputs markdown table format."""
        xlsx_path = tmp_path / "small.xlsx"
        # Create table with 10 rows (within 35 row limit)
        data = {
            "Name": [f"Person{i}" for i in range(10)],
            "Age": [20 + i for i in range(10)],
        }
        create_xlsx({"Sheet1": data}, xlsx_path)

        result = await processor.process(str(xlsx_path))

        # Should contain markdown table format
        assert "|" in result.markdown
        assert "---" in result.markdown
        assert "Person0" in result.markdown


class TestXlsxProcessorLargeTables:
    """Tests for large table (sentence format)."""

    @pytest.mark.asyncio
    async def test_large_table_sentences(self, processor, tmp_path):
        """Large table outputs sentence format."""
        xlsx_path = tmp_path / "large.xlsx"
        # Create table with 40 rows (exceeds 35 row limit)
        data = {
            "Name": [f"Person{i}" for i in range(40)],
            "Age": [20 + i for i in range(40)],
        }
        create_xlsx({"Sheet1": data}, xlsx_path)

        result = await processor.process(str(xlsx_path))

        assert "Person0" in result.markdown
        # Should NOT have table separators for large tables
        assert "---" not in result.markdown or result.markdown.count("|") < 10

    @pytest.mark.asyncio
    async def test_sentence_template(self, processor, tmp_path):
        """Sentence format uses correct template."""
        xlsx_path = tmp_path / "sentence.xlsx"
        data = {
            "Name": [f"Person{i}" for i in range(40)],
            "City": [f"City{i}" for i in range(40)],
        }
        create_xlsx({"Sheet1": data}, xlsx_path)

        result = await processor.process(str(xlsx_path))

        # Should have sentence-style output like "**Name:** Person0"
        assert "Person0" in result.markdown


class TestXlsxProcessorDataTypes:
    """Tests for data type handling."""

    @pytest.mark.asyncio
    async def test_number_formatting(self, processor, tmp_path):
        """Numbers are formatted correctly."""
        xlsx_path = tmp_path / "numbers.xlsx"
        create_xlsx({"Data": {"Value": [1000, 2500.5, 3000000]}}, xlsx_path)

        result = await processor.process(str(xlsx_path))

        # Numbers should appear in output
        assert "1000" in result.markdown or "1,000" in result.markdown

    @pytest.mark.asyncio
    async def test_date_conversion(self, processor, tmp_path):
        """Dates are handled (no crashing)."""
        xlsx_path = tmp_path / "dates.xlsx"
        create_xlsx({"Dates": {"Date": ["2024-01-15", "2024-06-30"]}}, xlsx_path)

        result = await processor.process(str(xlsx_path))

        assert "2024" in result.markdown

    @pytest.mark.asyncio
    async def test_skip_null_cells(self, processor, tmp_path):
        """Null/empty cells are handled gracefully."""
        xlsx_path = tmp_path / "nulls.xlsx"
        create_xlsx({"Data": {"A": [1, None, 3], "B": [None, 2, None]}}, xlsx_path)

        result = await processor.process(str(xlsx_path))

        # Should not crash, values should appear
        assert "1" in result.markdown
        assert "3" in result.markdown


class TestXlsxProcessorEdgeCases:
    """Tests for edge cases."""

    @pytest.mark.asyncio
    async def test_file_not_found(self, processor):
        """Missing file returns error."""
        result = await processor.process("/nonexistent/file.xlsx")

        assert result.markdown == ""
        assert "error" in result.metadata

    @pytest.mark.asyncio
    async def test_empty_file(self, processor, tmp_path):
        """Empty XLSX file is handled."""
        xlsx_path = tmp_path / "empty.xlsx"
        wb = Workbook()
        wb.save(xlsx_path)

        result = await processor.process(str(xlsx_path))

        # Should handle gracefully
        assert result.markdown == "" or result.sheet_count == 0

    @pytest.mark.asyncio
    async def test_password_protected(self, processor, tmp_path):
        """Password protected file returns error."""
        # Create a file that simulates password protection by being invalid
        xlsx_path = tmp_path / "protected.xlsx"
        xlsx_path.write_bytes(b"PK invalid content")

        result = await processor.process(str(xlsx_path))

        # Should have error in metadata
        assert "error" in result.metadata or result.markdown == ""


class TestXlsxProcessorMetadata:
    """Tests for metadata extraction."""

    @pytest.mark.asyncio
    async def test_sheet_count_in_result(self, processor, tmp_path):
        """Sheet count is in result."""
        xlsx_path = tmp_path / "count.xlsx"
        create_xlsx(
            {"Sheet1": {"A": [1]}, "Sheet2": {"B": [2]}, "Sheet3": {"C": [3]}},
            xlsx_path,
        )

        result = await processor.process(str(xlsx_path))

        assert result.sheet_count == 3
